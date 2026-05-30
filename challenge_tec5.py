# -*- coding: utf-8 -*-
"""
Challenge Integrador - GuardiánClima ITBA
Sistema unificado: Login, Registro, APIs de Clima (OpenWeatherMap) y Gemini (IA).
Módulo de exportación avanzada a formato Excel integrado.
"""

import os
from collections import Counter
import requests
import csv
import json
from datetime import datetime
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

# Cargar las variables de entorno desde el archivo .env oculto
load_dotenv()

# Clase para el manejo de estilos de color en la terminal mediante códigos ANSI
class Color:
    VERDE = '\033[92m'
    ROJO = '\033[91m'
    AMARILLO = '\033[93m'
    AZUL = '\033[94m'
    CELESTE = '\033[96m'
    MAGENTA = '\033[95m'
    BOLD = '\033[1m'
    FIN = '\033[0m'

USERS_FILE = 'usuarios_simulados.csv'

# Extracción segura de las credenciales
OWM_API_KEY = os.getenv("OPENWEATHER_API_KEY")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# Validación preventiva de seguridad
if not OWM_API_KEY or not GEMINI_API_KEY:
    print(f"{Color.ROJO}{Color.BOLD}ERROR CRITICO: No se encontraron las API Keys.{Color.FIN}")
    print(f"{Color.AMARILLO}Asegurate de haber creado el archivo .env con OPENWEATHER_API_KEY y GEMINI_API_KEY.{Color.FIN}")
    exit()
# =====================================================================
# MÓDULO DE FUINCIONES PARA USUARIOS
# =====================================================================

def save_user_credentials(username, password):
    """
    Guarda las credenciales de un nuevo usuario en un archivo CSV.
    Nota de seguridad: Guardar contraseñas en texto plano es una práctica
    insegura. Se utiliza en este entorno únicamente con fines educativos
    para la simulación del acceso.
    """
    file_exists = os.path.isfile(USERS_FILE)
    with open(USERS_FILE, 'a', newline='') as csvfile:
        fieldnames = ['usuario', 'contraseña']
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        if not file_exists:
            writer.writeheader() 

        writer.writerow({'usuario': username, 'contraseña': password})
    print(f"{Color.VERDE}Credenciales para '{username}' guardadas en {USERS_FILE}{Color.FIN}")

def load_users():
    """
    Lee el archivo CSV de usuarios y carga las credenciales en un diccionario
    en memoria para realizar las validaciones de forma más rápida.
    """
    users = {}
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                users[row['usuario']] = row['contraseña']
    return users

def username_exists(username):
    """
    Comprueba si un nombre de usuario ya se encuentra registrado en el archivo.
    """
    users = load_users()
    return username in users

def verify_credentials(username, password):
    """
    Verifica si la contraseña ingresada coincide con el registro del usuario.
    """
    users = load_users()
    return users.get(username) == password

def validar_password(password):
    """
    Valida que la contraseña cumpla con los requisitos mínimos de seguridad.
    Devuelve una lista con los errores detectados para informar al usuario.
    """
    errores = []

    if len(password) < 12:
        errores.append("tener al menos 12 caracteres")

    if not any(char.isdigit() for char in password):
        errores.append("contener al menos un numero")

    if not any(char.isupper() for char in password):
        errores.append("contener al menos una mayuscula")

    return errores

# =====================================================================
# MÓDULO DE INTEGRACIÓN DE APIs (Clima e IA) y EXPORTACIÓN
# =====================================================================

def consultar_clima(ciudad: str, api_key: str) -> dict:
    """
    Consulta la API de OpenWeatherMap mediante una petición GET.
    Filtra los datos recibidos y muestra la información en la consola.
    """
    base_url = "https://api.openweathermap.org/data/2.5/weather"
    parametros = {
        'q': ciudad,
        'appid': api_key,
        'units': 'metric',
        'lang': 'es'
    }

    try:
        response = requests.get(base_url, params=parametros, timeout=10)
        response.raise_for_status()
        data = response.json()

        clima_filtrado = {
            "temperatura": data["main"]["temp"],
            "sensacion_termica": data["main"]["feels_like"],
            "humedad": data["main"]["humidity"],
            "condicion": data["weather"][0]["description"],
            "viento": data["wind"]["speed"]
        }

        print(f"\n{Color.CELESTE}{Color.BOLD}--- Clima Actual en {ciudad.capitalize()} ---{Color.FIN}")
        print(f"{Color.CELESTE}Temperatura:{Color.FIN} {clima_filtrado['temperatura']}°C (Sensacion termica: {clima_filtrado['sensacion_termica']}°C)")
        print(f"{Color.CELESTE}Condicion:{Color.FIN} {clima_filtrado['condicion'].capitalize()}")
        print(f"{Color.CELESTE}Humedad:{Color.FIN} {clima_filtrado['humedad']}%")
        print(f"{Color.CELESTE}Viento:{Color.FIN} {clima_filtrado['viento']} km/h")
        print(Color.CELESTE + "-" * 40 + Color.FIN)

        return clima_filtrado

    except requests.exceptions.HTTPError as errh:
        if response.status_code == 401:
            print(f"{Color.ROJO}Error OWM: API Key invalida o no activada.{Color.FIN}")
        elif response.status_code == 404:
            print(f"{Color.ROJO}Error OWM: La ciudad '{ciudad}' no fue encontrada.{Color.FIN}")
        else:
            print(f"{Color.ROJO}Error HTTP OWM: {errh}{Color.FIN}")
    except requests.exceptions.RequestException as err:
        print(f"{Color.ROJO}Error de conexion/peticion OWM: {err}{Color.FIN}")
    except json.JSONDecodeError:
        print(f"{Color.ROJO}Error OWM: La respuesta de la API no es un JSON valido.{Color.FIN}")
    except KeyError as e:
        print(f"{Color.ROJO}Error estructural en la API. Falta el dato: {e}{Color.FIN}")

    return None

def guardar_historial_clima(usuario: str, ciudad: str, datos_clima: dict) -> None:
    """
    Guarda los datos de la consulta climática en el archivo historial_global.csv.
    Si el archivo no existe, escribe los encabezados correspondientes.
    """
    if not datos_clima:
        return

    archivo_csv = "historial_global.csv"
    archivo_existe = os.path.isfile(archivo_csv)
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    fila_historial = [
        usuario,
        ciudad.capitalize(),
        fecha_hora,
        datos_clima["temperatura"],
        datos_clima["condicion"],
        datos_clima["humedad"],
        datos_clima["viento"]
    ]

    try:
        with open(archivo_csv, mode="a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            if not archivo_existe:
                writer.writerow(["NombreDeUsuario", "Ciudad", "Fecha/Hora", "Temperatura_C", "Condicion_Clima", "Humedad_Porcentaje", "Viento_kmh"])
            writer.writerow(fila_historial)
            print(f"{Color.VERDE}[*] Consulta guardada exitosamente en el historial_global.csv.{Color.FIN}")
    except IOError as e:
        print(f"{Color.ROJO}Error critico al intentar escribir en el archivo CSV: {e}{Color.FIN}")

def obtener_consejo_vestimenta(datos_clima: dict, api_key: str) -> None:
    """
    Envia los datos meteorologicos de la ultima consulta realizada al modelo
    Gemini para obtener una recomendacion sobre que ropa vestir.
    """
    if not datos_clima:
        print(f"{Color.AMARILLO}No hay datos climaticos previos. Por favor, consulta el clima primero (Opcion 1).{Color.FIN}")
        return

    genai.configure(api_key=api_key)

    prompt_sistema = (
        f"Eres un experto asesor de imagen y supervivencia urbana. "
        f"Actualmente la temperatura es de {datos_clima['temperatura']}°C "
        f"(sensacion termica de {datos_clima['sensacion_termica']}°C), "
        f"con humedad del {datos_clima['humedad']}% y vientos de {datos_clima['viento']} km/h. "
        f"El cielo presenta: {datos_clima['condicion']}. "
        f"Dame un consejo practico, directo y muy breve (maximo 3 lineas) sobre como debo vestirme "
        f"y si necesito llevar algun accesorio adicional (paraguas, abrigo, lentes)."
    )

    try:
        model = genai.GenerativeModel('gemini-2.5-flash')
        print(f"\n{Color.MAGENTA}Consultando a la Inteligencia Artificial (Gemini)...{Color.FIN}")
        response = model.generate_content(prompt_sistema)

        if response.text:
            print(f"\n{Color.MAGENTA}{Color.BOLD}[Consejo de Vestimenta IA]{Color.FIN}")
            print(f"{response.text}\n")
        else:
            print(f"{Color.ROJO}La IA no devolvio un consejo valido.{Color.FIN}")
            if response.prompt_feedback:
                print(f"{Color.ROJO}Filtro de seguridad: {response.prompt_feedback}{Color.FIN}")
    except Exception as e:
        print(f"{Color.ROJO}Error al contactar la API de Gemini: {e}{Color.FIN}")

def ver_historial_personal(usuario: str, ciudad: str) -> None:
    """
    Filtra el archivo CSV global para extraer y mostrar de forma tabulada
    las consultas realizadas exclusivamente por el usuario autenticado.
    """
    archivo_csv = "historial_global.csv"

    if not os.path.exists(archivo_csv):
        print(f"{Color.AMARILLO}El historial global aun no existe. Realiza una consulta de clima primero (Opcion 1).{Color.FIN}")
        return

    registros_encontrados = []

    try:
        with open(archivo_csv, mode="r", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                if row["NombreDeUsuario"] == usuario and row["Ciudad"].lower() == ciudad.lower():
                    registros_encontrados.append(row)

        if not registros_encontrados:
            print(f"\n{Color.AMARILLO}No se encontraron consultas previas de la ciudad '{ciudad.capitalize()}' para tu usuario.{Color.FIN}")
        else:
            print(f"\n{Color.AZUL}{Color.BOLD}--- Historial Personal: {ciudad.capitalize()} (Usuario: {usuario}) ---{Color.FIN}")
            print(f"{Color.AZUL}{'Fecha/Hora':<22} | {'Temp (°C)':<10} | {'Condicion':<18} | {'Humedad (%)':<12} | {'Viento (km/h)'}{Color.FIN}")
            print(Color.AZUL + "-" * 85 + Color.FIN)

            for reg in registros_encontrados:
                print(f"{reg['Fecha/Hora']:<22} | {reg['Temperatura_C']:<10} | {reg['Condicion_Clima'].capitalize():<18} | {reg['Humedad_Porcentaje']:<12} | {reg['Viento_kmh']}")
            print(Color.AZUL + "-" * 85 + Color.FIN)

    except IOError as e:
        print(f"{Color.ROJO}Error al intentar leer el archivo CSV: {e}{Color.FIN}")
    except KeyError as e:
        print(f"{Color.ROJO}Error estructural en el archivo CSV. Falta la columna: {e}{Color.FIN}")

def exportar_excel_estilizado():
    """
    Lee los datos del archivo CSV e invoca las herramientas de la libreria
    openpyxl para estructurar un libro Excel (.xlsx) formateado con
    estilos visuales, dimensiones calculadas, autofiltros y paneles estaticos.
    """
    archivo_csv = "historial_global.csv"
    archivo_excel = "GuardianClima_Reporte.xlsx"

    if not os.path.exists(archivo_csv):
        print(f"{Color.AMARILLO}No se puede exportar: El archivo {archivo_csv} aun no existe.{Color.FIN}")
        return

    print(f"\n{Color.CELESTE}Generando archivo Excel con formato avanzado...{Color.FIN}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos Historicos"

    # Definicion de colores y fuentes para cumplir con el diseño profesional
    fondo_encabezado = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fondo_cebra = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")
    fuente_encabezado = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fuente_cuerpo = Font(name="Arial", size=10)
    borde_fino = Border(
        left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9")
    )
    alineacion_centro = Alignment(horizontal="center", vertical="center")

    try:
        with open(archivo_csv, mode="r", encoding="utf-8") as file:
            reader = csv.reader(file)
            datos = list(reader)

            for row_idx, fila in enumerate(datos, start=1):
                ws.append(fila)
                
                for col_idx in range(1, len(fila) + 1):
                    celda = ws.cell(row=row_idx, column=col_idx)
                    celda.border = borde_fino
                    
                    if row_idx == 1:
                        celda.fill = fondo_encabezado
                        celda.font = fuente_encabezado
                        celda.alignment = alineacion_centro
                    else:
                        celda.font = fuente_cuerpo
                        if row_idx % 2 == 0:
                            celda.fill = fondo_cebra

            # Ajuste automatico de las columnas segun la longitud del contenido
            for col in ws.columns:
                max_length = 0
                col_letra = col[0].column_letter
                for celda in col:
                    if celda.value:
                        max_length = max(max_length, len(str(celda.value)))
                ws.column_dimensions[col_letra].width = max_length + 2

            # Habilitacion de filtros e inmovilizacion de la primera fila
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            wb.save(archivo_excel)
            print(f"{Color.VERDE}[✓] Reporte Excel generado de forma exitosa: {archivo_excel}{Color.FIN}")

    except Exception as e:
        print(f"{Color.ROJO}Error al generar el archivo Excel: {e}{Color.FIN}")

def mostrar_estadisticas_globales() -> None:
    """
    Procesa todas las entradas del CSV para calcular metricas del sistema
    (total de consultas, ciudad mas buscada y temperatura promedio global).
    """
    archivo_csv = "historial_global.csv"

    if not os.path.exists(archivo_csv):
        print(f"{Color.AMARILLO}El historial global aun no existe. Nadie ha realizado consultas de clima todavia.{Color.FIN}")
        return

    total_consultas = 0
    ciudades_consultadas = []
    suma_temperaturas = 0.0

    try:
        with open(archivo_csv, mode="r", encoding="utf-8") as file:
            reader = csv.DictReader(file)

            for row in reader:
                total_consultas += 1
                ciudades_consultadas.append(row["Ciudad"])
                suma_temperaturas += float(row["Temperatura_C"])

        if total_consultas == 0:
            print(f"{Color.AMARILLO}El historial esta vacio. No hay datos suficientes para calcular estadisticas.{Color.FIN}")
            return

        ciudad_mas_frecuente = Counter(ciudades_consultadas).most_common(1)[0][0]
        temperatura_promedio = suma_temperaturas / total_consultas

        print(f"\n{Color.VERDE}" + "="*50)
        print(f"  ESTADISTICAS GLOBALES DE GUARDIANCLIMA ")
        print("="*50 + Color.FIN)
        print(f" Total de consultas en el sistema: {total_consultas}")
        print(f" Ciudad mas consultada:          {ciudad_mas_frecuente}")
        print(f" Temperatura promedio global:    {temperatura_promedio:.2f} °C")
        print(f"{Color.VERDE}" + "="*50 + Color.FIN)

        print(f"\n{Color.AMARILLO}{Color.BOLD}NOTA PARA ANALISIS EXTERNO:{Color.FIN}")
        print("Recuerda que puedes abrir el archivo 'historial_global.csv' en Excel o Google Sheets")
        print("para generar tus visualizaciones (graficos de barras, lineas y torta).")

        exportar = input(f"\n{Color.CELESTE}¿Deseas exportar estos datos estructurados a un archivo Excel (.xlsx)? (S/N): {Color.FIN}")
        if exportar.strip().upper() == 'S':
            exportar_excel_estilizado()

    except IOError as e:
        print(f"{Color.ROJO}Error al intentar leer el archivo CSV: {e}{Color.FIN}")
    except ValueError as e:
        print(f"{Color.ROJO}Error de formato de datos (se esperaba un numero en la temperatura): {e}{Color.FIN}")
    except KeyError as e:
        print(f"{Color.ROJO}Error estructural en el archivo CSV. Faltan encabezados: {e}{Color.FIN}")

def mostrar_acerca_de() -> None:
    """
    Imprime la seccion informativa que detalla el funcionamiento interno de la aplicacion,
    los flujos de los datos, las advertencias de seguridad y la fundamentacion tecnica
    del nuevo modulo de generacion de reportes Excel.
    """
    R = Color.ROJO
    G = Color.VERDE
    Y = Color.AMARILLO
    B = Color.AZUL
    C = Color.CELESTE
    W = Color.FIN
    BOLD = Color.BOLD

    print(C + BOLD + r"""
  _____                      _ _             _____ _ _
 / ____|                    | (_)           / ____| (_)
| |  __ _   _  __ _ _ __  __| |_  __ _ _ __| |    | |_ _ __ ___   __ _
| | |_ | | | |/ _` | '__|/ _` | |/ _` | '_ \ |    | | | '_ ` _ \ / _` |
| |__| | |_| | (_| | |  | (_| | | (_| | | | | |____| | | | | | | | (_| |
 \_____|\\__,_|\\__,_|_|   \\__,_|_|\\__,_|_| |_|\\_____|_|_|_| |_| |_|\\__,_|
    """ + W)

    print(B + "=" * 80 + W)
    print(C + BOLD + "           SISTEMA INTEGRADOR DE GESTION Y MONITOREO CLIMATICO" + W)
    print(B + "=" * 80 + W)

    print(BOLD + "\n DESCRIPCION GENERAL" + W)
    print(
        "GuardianClima ITBA es una solucion de consola desarrollada en Python que unifica\n"
        "el consumo de servicios web (APIs REST), persistencia local en CSV y generacion\n"
        "de contenido mediante inteligencia artificial."
    )

    print(Y + BOLD + "\n GUIA DE USO Y FLUJOS" + W)
    print(BOLD + " Flujo Pre-Login:" + W + " Verificacion y registro con reglas de validacion de contraseñas.")
    print(BOLD + " Flujo Post-Login:" + W + " Consumo de OpenWeatherMap, filtrado de datos, analisis\n"
          "                   estadistico y sugerencias usando Google Gemini.")

    print(R + BOLD + "\n ARQUITECTURA Y CIBERSEGURIDAD" + W)
    print(" Gestion de Usuarios: En este TP guardamos credenciales en texto plano en el CSV.")
    print(" " + BOLD + "ADVERTENCIA CRITICA:" + W + " En produccion, esto debe ser reemplazado por funciones hash\n"
          " (como SHA-256 o bcrypt) para proteger los datos frente a vulnerabilidades.")

    # Fundamentación técnica del módulo Excel requerida por la consigna
    print(G + BOLD + "\n FUNDAMENTACION: MODULO DE EXPORTACION EXCEL (.XLSX)" + W)
    print(" El sistema separa el almacenamiento transaccional (CSV) de la capa de analisis\n"
          " y negocio (BI). La integracion de la libreria `openpyxl` fundamenta un proceso de\n"
          " automatizacion de datos (Data Automation) que mitiga errores de formato humanos.\n"
          " Mediante codigo, se genera una estructura XML nativa comprimida que pre-formatea\n"
          " la hoja con estilos corporativos, auto-ajuste de columnas, paneles fijos para\n"
          " grandes volumenes de filas y filtros activos, optimizando la posterior toma de\n"
          " decisiones y el desarrollo de graficos estadisticos complejos.")

    print(B + "\n" + "=" * 80 + W)
    print(R + BOLD + "                        EQUIPO DESARROLLADOR" + W)
    print(B + "=" * 80 + W)

    print(R + """      .---.       """ + W + BOLD + " GRUPO:      " + W + R + BOLD + "EQUIPO DINAMITA" + W)
    print(R + """     (     )      """ + W + BOLD + " COMISION:   " + W + "Comisión U" + W)
    print(R + """      \\   /       """ + BOLD + " INTEGRANTES:" + W)
    print(R + """  .----'-'----.   """ + W + "  • Tomas Hauser          " + BOLD + "Legajo: " + W + "69193")
    print(R + """ |  """ + Y + BOLD + "DINAMITA" + R + """  |   """ + W + "  • Lautaro Kilimik       " + BOLD + "Legajo: " + W + "68951")
    print(R + """ | """ + B + BOLD + "COMISION U" + R + """ |   """ + W + "  • Ignacio Frontera      " + BOLD + "Legajo: " + W + "68767")
    print(R + """  '------------'  """ + W + "  • Ignacio Herrera       " + BOLD + "Legajo: " + Y + "68674" + W)
    print(B + "=" * 80 + W)


# =====================================================================
# FLUJO PRINCIPAL DEL PROGRAMA
# =====================================================================

if __name__ == "__main__":
    usuario_actual = None 
    ultima_consulta_clima = None 

    while True:
        # Menú de Acceso (Usuario sin autenticar)
        if usuario_actual is None:
            print(f"\n{Color.AZUL}{Color.BOLD}=== Menu de Acceso ==={Color.FIN}")
            print("1. Iniciar Sesion")
            print("2. Registrarse")
            print("3. Salir")

            choice = input(f"{Color.AMARILLO}Ingrese su opcion: {Color.FIN}")

            if choice == '1': 
                print(f"\n{Color.AZUL}--- Iniciar Sesion ---{Color.FIN}")
                while True:
                    print(f"{Color.AMARILLO}Presione Enter con la celda vacia para volver al menu.{Color.FIN}")
                    username = input("Usuario: ")
                    if username == "":
                      break
                    if not username_exists(username):
                        print(f"{Color.ROJO}Error: Usuario no registrado. Por favor, registrese.\n{Color.FIN}")
                    else:
                        password = input("Contraseña: ")
                        if verify_credentials(username, password):
                            print(f"{Color.VERDE}¡Bienvenido, {username}! Ha iniciado sesion correctamente.\n{Color.FIN}")
                            usuario_actual = username 
                            break
                        else:
                            print(f"{Color.ROJO}Error: Contraseña incorrecta. Intente de nuevo.\n{Color.FIN}")

            elif choice == '2':  
                print(f"\n{Color.AZUL}--- Registrarse ---{Color.FIN}")

                while True:
                    print(f"{Color.AMARILLO}Presione Enter con la celda vacia para volver al menu.{Color.FIN}")
                    new_username = input("Ingrese un nuevo usuario: ")

                    if new_username == "":
                        break

                    if username_exists(new_username):
                        print(f"{Color.ROJO}Error: El nombre de usuario ya existe. Elija otro.\n{Color.FIN}")
                        continue
                    else:
                        break

                if new_username == "":
                    continue

                while True:
                    new_password = input("Ingrese una nueva contraseña: ")

                    if new_password == "":
                        break

                    errores = validar_password(new_password)

                    if errores:
                        print(f"{Color.ROJO}Errores encontrados: La contraseña debe {', '.join(errores)}.{Color.FIN}")
                        print(f"{Color.AMARILLO}Sugerencia: Intente mezclar palabras, numeros y caracteres especiales.\n{Color.FIN}")
                        continue
                    else:
                        break

                if new_password == "":
                    continue

                save_user_credentials(new_username, new_password)
                print(f"{Color.VERDE}Usuario '{new_username}' registrado exitosamente.\n{Color.FIN}")
                usuario_actual = new_username

            elif choice == '3': 
                print(f"{Color.CELESTE}Saliendo del sistema. ¡Hasta luego!{Color.FIN}")
                break

            else:
                print(f"{Color.ROJO}Opcion invalida. Intente de nuevo.\n{Color.FIN}")

        # Menú Principal (Usuario autenticado)
        else:
            print(f"\n{Color.AZUL}{Color.BOLD}=== Menu Principal (Usuario: {usuario_actual}) ==={Color.FIN}")
            print("1. Consultar Clima Actual y Guardar en Historial")
            print("2. Ver Mi Historial Personal")
            print("3. Estadisticas Globales / Exportar Excel")
            print("4. Consejo IA: ¿Como Me Visto Hoy?")
            print("5. Acerca De...")
            print("6. Cerrar Sesion")

            choice = input(f"{Color.AMARILLO}Ingrese su opcion: {Color.FIN}")

            if choice == '1':
                ciudad_input = input("Ingresa el nombre de la ciudad: ")
                datos = consultar_clima(ciudad_input, OWM_API_KEY)

                if datos:
                    guardar_historial_clima(usuario_actual, ciudad_input, datos)
                    ultima_consulta_clima = datos 
                
                input(f"\n{Color.AMARILLO}{Color.BOLD}[Presione Enter para continuar...]{Color.FIN}")
                
            elif choice == '2':
                print("\n--- Ver Mi Historial Personal ---")
                ciudad_buscar = input(f"Ingresa la ciudad que deseas buscar en tu historial: {Color.FIN}")
                ver_historial_personal(usuario_actual, ciudad_buscar)
                input(f"\n{Color.AMARILLO}{Color.BOLD}[Presione Enter para continuar...]{Color.FIN}")
                
            elif choice == '3':
                mostrar_estadisticas_globales()
                input(f"\n{Color.AMARILLO}{Color.BOLD}[Presione Enter para continuar...]{Color.FIN}")
                
            elif choice == '4':
                obtener_consejo_vestimenta(ultima_consulta_clima, GEMINI_API_KEY)
                input(f"\n{Color.AMARILLO}{Color.BOLD}[Presione Enter para continuar...]{Color.FIN}")
                
            elif choice == '5':
                mostrar_acerca_de()
                input(f"\n{Color.AMARILLO}{Color.BOLD}[Presione Enter para continuar...]{Color.FIN}")
                
            elif choice == '6':
                print(f"{Color.CELESTE}Cerrando sesion de {usuario_actual}...{Color.FIN}")
                usuario_actual = None 
                ultima_consulta_clima = None
            else:
                print(f"{Color.ROJO}Opcion invalida. Intente de nuevo.\n{Color.FIN}")
